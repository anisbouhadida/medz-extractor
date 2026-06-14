#!/usr/bin/env python3
# SPDX-FileCopyrightText: 2026 Anis Bouhadida
# SPDX-License-Identifier: AGPL-3.0-or-later
"""Extract Algerian pharmaceutical nomenclature Excel files to CSV.

Usage:
    python3 scripts/extract_medz.py input output
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import shutil
import sys
import time
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

EXPECTED_SHEETS = {
    "nomenclature": "nomenclature.csv",
    "non renouveles": "non_renouveles.csv",
    "retraits": "retraits.csv",
}
CSV_FILENAMES = tuple(EXPECTED_SHEETS.values())

NOMENCLATURE_THRESHOLD = 8
DEFAULT_THRESHOLD = 6
FOOTER_PREFIXES = ("F=", "I=", "Nb:")
FOOTER_MAX_NON_EMPTY = 3
INPUT_NAME_PATTERN = re.compile(r"^\d{4}-\d{2}\.xlsx$")

logger = logging.getLogger(__name__)


class ExtractionError(ValueError):
    """Raised for expected extraction failures with user-facing messages."""


def utc_timestamp() -> str:
    """Return the archive timestamp in YYYYMMDDTHHMMSSZ format."""
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def normalize_sheet_name(name: str) -> str:
    """Normalize a sheet name for accent/case/spacing tolerant matching."""
    decomposed = unicodedata.normalize("NFD", name)
    stripped = "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")
    lowered = stripped.lower()
    return re.sub(r"[\s_\-]+", " ", lowered).strip()


def detect_sheets(sheet_names: list[str]) -> dict[str, str]:
    """Map original workbook sheet names to required output CSV filenames."""
    normalized_actuals = {normalize_sheet_name(sn): sn for sn in sheet_names}
    matched: dict[str, str] = {}
    missing: list[str] = []

    for canonical, csv_name in EXPECTED_SHEETS.items():
        found = False
        for norm_actual, original in normalized_actuals.items():
            if norm_actual == canonical or norm_actual.startswith(canonical + " "):
                matched[original] = csv_name
                logger.info("Sheet '%s' matched canonical '%s'.", original, canonical)
                found = True
                break
        if not found:
            missing.append(canonical)

    if missing:
        raise ExtractionError(
            "Missing expected sheet(s): "
            + ", ".join(f"'{m}'" for m in missing)
            + ". Available sheets: "
            + ", ".join(f"'{s}'" for s in sheet_names)
            + "."
        )

    return matched


def _cell_to_str(value: object) -> str:
    """Convert a cell value to a stripped, single-line string."""
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text


def count_non_empty_cells(row: tuple[object, ...]) -> int:
    """Count non-empty cells in a worksheet row."""
    return sum(1 for cell in row if _cell_to_str(cell) != "")


def is_tabular_row(row: tuple[object, ...], threshold: int) -> bool:
    """Return whether a row has enough populated cells to be tabular."""
    return count_non_empty_cells(row) >= threshold


def is_footer_row(row: tuple[object, ...]) -> bool:
    """Return whether a sparse row is a footer legend row."""
    if count_non_empty_cells(row) > FOOTER_MAX_NON_EMPTY:
        return False
    for cell in row:
        text = _cell_to_str(cell)
        if text and any(text.startswith(prefix) for prefix in FOOTER_PREFIXES):
            return True
    return False


def detect_header_row(rows: list[tuple[object, ...]], threshold: int) -> int:
    """Find the first tabular row followed by another tabular row."""
    for i in range(len(rows) - 1):
        if is_tabular_row(rows[i], threshold) and is_tabular_row(
            rows[i + 1],
            threshold,
        ):
            logger.info("Header row detected at sheet row index %d.", i)
            return i

    raise ExtractionError(
        "Header row not found: no row with >= "
        f"{threshold} non-empty cells followed by another tabular row."
    )


def extract_data(
    rows: list[tuple[object, ...]],
    header_index: int,
) -> tuple[list[str], list[list[str]]]:
    """Extract header and data rows, stopping before footer content."""
    header = [_cell_to_str(c) for c in rows[header_index]]
    data_rows: list[list[str]] = []

    i = header_index + 1
    while i < len(rows):
        row = rows[i]
        if is_footer_row(row):
            logger.info("Footer marker detected at row index %d.", i)
            break
        if count_non_empty_cells(row) == 0:
            has_more_tabular = False
            for j in range(i + 1, len(rows)):
                if count_non_empty_cells(rows[j]) == 0:
                    continue
                if is_footer_row(rows[j]):
                    break
                if count_non_empty_cells(rows[j]) >= 2:
                    has_more_tabular = True
                break
            if not has_more_tabular:
                logger.info("Structural collapse detected at row index %d.", i)
                break
            i += 1
            continue

        data_rows.append([_cell_to_str(c) for c in row])
        i += 1

    if not data_rows:
        raise ExtractionError(
            "Extracted data has 0 rows after removing header and footer blocks."
        )

    logger.info("Extracted %d data rows.", len(data_rows))
    return header, data_rows


def parse_sheet(
    ws: object,
    sheet_label: str,
    threshold: int | None = None,
) -> tuple[list[str], list[list[str]]]:
    """Parse one worksheet into headers and rows."""
    if threshold is None:
        threshold = (
            NOMENCLATURE_THRESHOLD
            if "nomenclature" in sheet_label.lower()
            else DEFAULT_THRESHOLD
        )

    rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
    if not rows:
        raise ExtractionError(f"Sheet '{sheet_label}' is empty: no rows found.")

    header_index = detect_header_row(rows, threshold)
    logger.info("Sheet '%s': header row index %d.", sheet_label, header_index)
    return extract_data(rows, header_index)


def find_empty_columns(headers: list[str], data_rows: list[list[str]]) -> set[int]:
    """Identify columns where every row value is empty."""
    empty = set(range(len(headers)))

    for row in data_rows:
        for col_idx in list(empty):
            if col_idx < len(row) and row[col_idx] != "":
                empty.discard(col_idx)
        if not empty:
            break

    return empty


def drop_empty_columns(
    headers: list[str],
    data_rows: list[list[str]],
) -> tuple[list[str], list[list[str]]]:
    """Remove columns that are empty across all data rows."""
    empty_indices = find_empty_columns(headers, data_rows)
    dropped_headers = [headers[i] for i in sorted(empty_indices)]
    logger.info(
        "Dropping %d entirely-empty column(s): %s.",
        len(empty_indices),
        dropped_headers,
    )

    keep = [i for i in range(len(headers)) if i not in empty_indices]
    cleaned_headers = [headers[i] for i in keep]
    cleaned_rows = [[row[i] if i < len(row) else "" for i in keep] for row in data_rows]
    return cleaned_headers, cleaned_rows


def write_csv(
    headers: list[str],
    data_rows: list[list[str]],
    output_path: Path,
) -> Path:
    """Write one UTF-8 comma-delimited CSV file."""
    if not headers:
        raise ExtractionError("Cannot write CSV: headers list is empty.")
    if not data_rows:
        raise ExtractionError("Cannot write CSV: data_rows list is empty.")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh, delimiter=",", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)
            writer.writerows(data_rows)
    except OSError as exc:
        raise ExtractionError(f"Failed to write CSV to '{output_path}': {exc}") from exc

    logger.info("Wrote %d data rows to '%s'.", len(data_rows), output_path)
    return output_path


def process_workbook(xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Extract one workbook into an output directory."""
    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        raise ExtractionError(f"Failed to open workbook '{xlsx_path}': {exc}") from exc

    try:
        sheet_map = detect_sheets(wb.sheetnames)
        logger.info("Detected sheets: %s", sheet_map)
        generated: list[Path] = []

        for sheet_name, csv_filename in sheet_map.items():
            logger.info("Processing sheet '%s' -> '%s'.", sheet_name, csv_filename)
            headers, data_rows = parse_sheet(wb[sheet_name], csv_filename)
            logger.info(
                "Sheet '%s': header row found, %d data rows extracted.",
                sheet_name,
                len(data_rows),
            )
            headers, data_rows = drop_empty_columns(headers, data_rows)
            generated.append(write_csv(headers, data_rows, output_dir / csv_filename))
    finally:
        wb.close()

    return generated


def discover_inputs(input_dir: Path) -> list[Path]:
    """Return sorted .xlsx inputs after validating all names."""
    if not input_dir.is_dir():
        raise ExtractionError(f"Input directory does not exist: {input_dir}")

    inputs = sorted(input_dir.glob("*.xlsx"))
    if not inputs:
        raise ExtractionError(f"No .xlsx files found in input directory: {input_dir}")

    invalid = [p.name for p in inputs if not INPUT_NAME_PATTERN.match(p.name)]
    if invalid:
        raise ExtractionError(
            "Invalid input filename(s): "
            + ", ".join(sorted(invalid))
            + ". Expected YYYY-MM.xlsx."
        )

    return inputs


def archive_current_outputs(
    month_output_dir: Path,
    archive_root: Path,
    month: str,
    timestamp: str,
) -> Path | None:
    """Move current CSV outputs for a month into the archive tree."""
    current_csvs = sorted(month_output_dir.glob("*.csv"))
    if not current_csvs:
        return None

    archive_dir = archive_root / month / timestamp
    try:
        archive_dir.mkdir(parents=True, exist_ok=False)
        for csv_path in current_csvs:
            shutil.move(str(csv_path), str(archive_dir / csv_path.name))
    except OSError as exc:
        raise ExtractionError(
            f"Failed to archive existing outputs for '{month}': {exc}"
        ) from exc

    logger.info("Archived existing outputs for '%s' to '%s'.", month, archive_dir)
    return archive_dir


def promote_staged_outputs(staging_dir: Path, month_output_dir: Path) -> None:
    """Move staged CSVs into the visible month output directory."""
    staged_csvs = sorted(staging_dir.glob("*.csv"))
    staged_names = sorted(p.name for p in staged_csvs)
    if staged_names != sorted(CSV_FILENAMES):
        raise ExtractionError(
            f"Staged output mismatch: expected {sorted(CSV_FILENAMES)}, got {staged_names}."
        )

    try:
        month_output_dir.mkdir(parents=True, exist_ok=True)
        for csv_path in staged_csvs:
            shutil.move(str(csv_path), str(month_output_dir / csv_path.name))
    except OSError as exc:
        raise ExtractionError(f"Failed to promote staged outputs: {exc}") from exc


def process_all(
    input_dir: Path | str,
    output_dir: Path | str,
    *,
    archive_root: Path | str | None = None,
    clock: Callable[[], str] = utc_timestamp,
) -> int:
    """Process all input workbooks into month folders under output_dir."""
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    archive_path = (
        output_path.parent / "archive" if archive_root is None else Path(archive_root)
    )
    staging_root = output_path / ".staging"

    logger.info("Input directory: %s", input_path)
    logger.info("Output directory: %s", output_path)
    logger.info("Archive directory: %s", archive_path)

    inputs = discover_inputs(input_path)
    processed = 0

    for xlsx_path in inputs:
        month = xlsx_path.stem
        timestamp = clock()
        staging_dir = staging_root / f"{month}-{timestamp}"
        month_output_dir = output_path / month
        logger.info("Processing workbook '%s'.", xlsx_path)

        try:
            if staging_dir.exists():
                raise ExtractionError(
                    f"Staging directory already exists: {staging_dir}"
                )
            generated = process_workbook(xlsx_path, staging_dir)
            if len(generated) != len(CSV_FILENAMES):
                raise ExtractionError(
                    f"Expected {len(CSV_FILENAMES)} CSV files, generated {len(generated)}."
                )
            archive_current_outputs(
                month_output_dir,
                archive_path,
                month,
                timestamp,
            )
            promote_staged_outputs(staging_dir, month_output_dir)
            processed += 1
        finally:
            if staging_dir.exists():
                shutil.rmtree(staging_dir)

    if staging_root.exists() and not any(staging_root.iterdir()):
        staging_root.rmdir()

    logger.info("Processed %d workbook(s).", processed)
    return processed


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse script arguments."""
    parser = argparse.ArgumentParser(
        description=(
            "Extract all YYYY-MM.xlsx files from an input directory into "
            "month-scoped CSV output folders."
        )
    )
    parser.add_argument(
        "input_dir", type=Path, help="Directory containing YYYY-MM.xlsx files."
    )
    parser.add_argument(
        "output_dir", type=Path, help="Directory where CSV outputs are written."
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """Script entry point."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    args = parse_args(sys.argv[1:] if argv is None else argv)
    start = time.monotonic()
    try:
        count = process_all(args.input_dir, args.output_dir)
    except ExtractionError as exc:
        logger.error("%s", exc)
        return 1

    logger.info("Finished in %.2f s.", time.monotonic() - start)
    logger.info("Ready for Spring Batch ingestion: %d month folder(s).", count)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
