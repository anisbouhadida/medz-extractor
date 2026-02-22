"""Full-pipeline regression tests using real MIPH Excel fixtures.

Each fixture is an official-format ``.xlsx`` file stored in
``tests/fixtures/``.  The expected (golden) CSV outputs are
committed in ``tests/golden/<fixture_name>/``.

The tests run the full extraction pipeline (open workbook →
detect sheets → parse → drop empty columns → write CSV) and
compare every output byte-for-byte against the golden files.

Adding a new fixture:
    1. Place the ``.xlsx`` under ``tests/fixtures/``.
    2. Run ``medz-extractor <file> --out tests/golden/<name>/``
       to generate golden CSVs.
    3. Review the golden CSVs for correctness.
    4. Commit both the fixture and the golden CSVs.
    5. The parametrised test picks it up automatically.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List, Tuple

import pytest
from openpyxl import load_workbook

from medz_extractor.parser import parse_sheet
from medz_extractor.schema import drop_empty_columns
from medz_extractor.sheet_detector import detect_sheets
from medz_extractor.writer import write_csv

# ── Path constants ──────────────────────────────────────────────

TESTS_DIR = Path(__file__).resolve().parent
FIXTURES_DIR = TESTS_DIR / "fixtures"
GOLDEN_DIR = TESTS_DIR / "golden"

# Expected CSV filenames produced by the pipeline.
CSV_FILENAMES: Tuple[str, ...] = (
    "nomenclature.csv",
    "non_renouveles.csv",
    "retraits.csv",
)


# ── Discover fixtures automatically ────────────────────────────


def _discover_fixtures() -> List[str]:
    """Return sorted list of fixture basenames (without .xlsx).

    Only fixtures that have a corresponding golden directory are
    included (guards against uncommitted goldens).
    """
    fixtures: List[str] = []
    for xlsx in sorted(FIXTURES_DIR.glob("*.xlsx")):
        name = xlsx.stem
        golden_dir = GOLDEN_DIR / name
        if golden_dir.is_dir():
            fixtures.append(name)
    return fixtures


FIXTURE_NAMES: List[str] = _discover_fixtures()


# ── Helpers ─────────────────────────────────────────────────────


def _read_csv(path: Path) -> Tuple[List[str], List[List[str]]]:
    """Read a CSV file and return (headers, data_rows)."""
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh, delimiter=",")
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _run_pipeline(
    xlsx_path: Path,
    output_dir: Path,
) -> List[Path]:
    """Execute the full extraction pipeline on *xlsx_path*.

    Mirrors the logic of the ``process`` CLI command without
    Typer so that tests do not depend on CLI wiring.

    Returns:
        List of written CSV paths.
    """
    wb = load_workbook(
        str(xlsx_path),
        read_only=True,
        data_only=True,
    )
    try:
        sheet_map = detect_sheets(wb.sheetnames)
        generated: List[Path] = []

        for sheet_name, csv_filename in sheet_map.items():
            ws = wb[sheet_name]
            headers, data_rows = parse_sheet(ws, csv_filename)
            headers, data_rows = drop_empty_columns(
                headers,
                data_rows,
            )
            csv_path = output_dir / csv_filename
            write_csv(headers, data_rows, csv_path)
            generated.append(csv_path)
    finally:
        wb.close()

    return generated


# ── Regression tests ────────────────────────────────────────────


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestFullPipelineRegression:
    """Compare pipeline output against golden CSVs for each fixture.

    Each fixture is parametrised so failures pinpoint the exact
    input file and sheet that diverged.
    """

    def test_pipeline_produces_all_csvs(
        self,
        fixture_name: str,
        tmp_path: Path,
    ) -> None:
        """Pipeline generates exactly the 3 expected CSV files."""
        xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
        generated = _run_pipeline(xlsx, tmp_path)
        actual_names = sorted(p.name for p in generated)
        expected_names = sorted(CSV_FILENAMES)
        assert actual_names == expected_names, (
            f"Fixture '{fixture_name}': expected CSVs "
            f"{expected_names}, got {actual_names}."
        )

    def test_nomenclature_matches_golden(
        self,
        fixture_name: str,
        tmp_path: Path,
    ) -> None:
        """nomenclature.csv matches the golden file exactly."""
        _assert_csv_matches_golden(
            fixture_name,
            "nomenclature.csv",
            tmp_path,
        )

    def test_non_renouveles_matches_golden(
        self,
        fixture_name: str,
        tmp_path: Path,
    ) -> None:
        """non_renouveles.csv matches the golden file exactly."""
        _assert_csv_matches_golden(
            fixture_name,
            "non_renouveles.csv",
            tmp_path,
        )

    def test_retraits_matches_golden(
        self,
        fixture_name: str,
        tmp_path: Path,
    ) -> None:
        """retraits.csv matches the golden file exactly."""
        _assert_csv_matches_golden(
            fixture_name,
            "retraits.csv",
            tmp_path,
        )


def _assert_csv_matches_golden(
    fixture_name: str,
    csv_filename: str,
    tmp_path: Path,
) -> None:
    """Run the pipeline and compare one CSV against its golden copy.

    Comparison is done at the parsed-CSV level (list of lists) so
    that line-ending differences do not cause false failures.
    """
    xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
    _run_pipeline(xlsx, tmp_path)

    actual_path = tmp_path / csv_filename
    golden_path = GOLDEN_DIR / fixture_name / csv_filename

    assert actual_path.exists(), (
        f"Pipeline did not produce '{csv_filename}' for fixture '{fixture_name}'."
    )
    assert golden_path.exists(), (
        f"Golden file missing: {golden_path}. Run the pipeline manually to create it."
    )

    actual_hdr, actual_rows = _read_csv(actual_path)
    golden_hdr, golden_rows = _read_csv(golden_path)

    assert actual_hdr == golden_hdr, (
        f"[{fixture_name}/{csv_filename}] Header mismatch.\n"
        f"  Expected: {golden_hdr}\n"
        f"  Got:      {actual_hdr}"
    )
    assert len(actual_rows) == len(golden_rows), (
        f"[{fixture_name}/{csv_filename}] Row count mismatch. "
        f"Expected {len(golden_rows)}, got {len(actual_rows)}."
    )

    for i, (actual_row, golden_row) in enumerate(zip(actual_rows, golden_rows)):
        assert actual_row == golden_row, (
            f"[{fixture_name}/{csv_filename}] "
            f"Row {i + 1} mismatch.\n"
            f"  Expected: {golden_row}\n"
            f"  Got:      {actual_row}"
        )


# ── Sanity checks on fixture discovery ──────────────────────────


class TestFixtureDiscovery:
    """Ensure we have a healthy test setup."""

    def test_at_least_one_fixture_found(self) -> None:
        """Guard against empty parametrisation."""
        assert len(FIXTURE_NAMES) > 0, (
            "No fixtures found. Place .xlsx files in "
            f"'{FIXTURES_DIR}' and golden CSVs in "
            f"'{GOLDEN_DIR}/<name>/'."
        )

    def test_all_golden_dirs_have_three_csvs(self) -> None:
        """Every golden directory contains exactly 3 CSVs."""
        for name in FIXTURE_NAMES:
            golden_dir = GOLDEN_DIR / name
            csv_files = sorted(f.name for f in golden_dir.glob("*.csv"))
            assert csv_files == sorted(CSV_FILENAMES), (
                f"Golden dir '{golden_dir}' has {csv_files}, "
                f"expected {sorted(CSV_FILENAMES)}."
            )
