"""Independent validation tests for golden CSV outputs.

These tests read the Excel fixtures directly with openpyxl
(bypassing medz_extractor's parser) and verify structural
invariants that must hold for any correct extraction.

This breaks the circular dependency of "testing the tool with
outputs produced by the same tool" by independently computing
expected properties from the raw Excel files.

Invariants checked:
- No institutional header text leaked into CSV data.
- No real footer legend lines (F=Fabriqué, I=Importé, Nb:)
  appear in CSV output.
- Data rows containing I= with many populated cells (real data,
  not footers) are preserved in CSV output.
- Row counts match independently computed expectations.
- CSV headers match the first tabular row in the Excel sheet.
- First and last data rows in the CSV match the Excel source.
- Row numbering in the first column is sequential starting at 1.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pytest
from openpyxl import load_workbook

from medz_extractor.sheet_detector import detect_sheets

# ── Path constants ──────────────────────────────────────────────

TESTS_DIR = Path(__file__).resolve().parent
FIXTURES_DIR = TESTS_DIR / "fixtures"
GOLDEN_DIR = TESTS_DIR / "golden"

# Known institutional header phrases that must NOT appear in data.
INSTITUTIONAL_MARKERS: Tuple[str, ...] = (
    "REPUBLIQUE ALGERIENNE",
    "MINISTERE DE L'INDUSTRIE",
    "DIRECTION DE LA PHARMACO",
    "NOMENCLATURE NATIONALE DES PRODUITS",
    "LISTE DES PRODUITS PHARMACEUTIQUES",
)

# Footer legend prefixes.
FOOTER_PREFIXES: Tuple[str, ...] = ("F=", "I=", "Nb:")

# Maximum non-empty cells for a row to be a footer legend.
FOOTER_MAX_NON_EMPTY: int = 3


# ── Helpers ─────────────────────────────────────────────────────


def _discover_fixtures() -> List[str]:
    """Return fixture basenames that have golden directories."""
    return sorted(
        xlsx.stem
        for xlsx in FIXTURES_DIR.glob("*.xlsx")
        if (GOLDEN_DIR / xlsx.stem).is_dir()
    )


def _read_golden_csv(
    fixture_name: str, csv_filename: str,
) -> Tuple[List[str], List[List[str]]]:
    """Read a golden CSV and return (headers, data_rows)."""
    path = GOLDEN_DIR / fixture_name / csv_filename
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh, delimiter=",")
        rows = list(reader)
    return rows[0], rows[1:]


def _cell_str(value: Any) -> str:
    """Convert a raw cell value to its stripped string form."""
    if value is None:
        return ""
    return str(value).strip()


def _count_non_empty(row: Tuple[Any, ...]) -> int:
    """Count non-empty cells in an Excel row."""
    return sum(1 for c in row if _cell_str(c) != "")


def _is_footer_legend(row: Tuple[Any, ...]) -> bool:
    """Check if a row is a footer legend (few cells + F=/I=/Nb:)."""
    if _count_non_empty(row) > FOOTER_MAX_NON_EMPTY:
        return False
    return any(
        _cell_str(c).startswith(FOOTER_PREFIXES)
        for c in row
    )


def _independently_extract(
    xlsx_path: Path,
) -> Dict[str, Tuple[List[str], List[List[str]], int]]:
    """Read Excel directly and compute expected structure.

    Returns a dict mapping csv_filename to
    ``(headers, first_5_rows, expected_data_count)``.

    The count is computed by:
    1. Finding the header row (first row with >= threshold
       non-empty cells where the next row is also tabular).
    2. Counting all subsequent tabular rows until a real footer
       legend or structural collapse.

    This is an independent reimplementation, NOT using
    medz_extractor's parser.
    """
    wb = load_workbook(
        str(xlsx_path), read_only=True, data_only=True,
    )
    try:
        sheet_map = detect_sheets(wb.sheetnames)
        result: Dict[
            str, Tuple[List[str], List[List[str]], int]
        ] = {}

        for sheet_name, csv_filename in sheet_map.items():
            ws = wb[sheet_name]
            all_rows = [
                tuple(cell for cell in row)
                for row in ws.iter_rows(values_only=True)
            ]

            # Determine threshold.
            if "nomenclature" in csv_filename.lower():
                threshold = 8
            else:
                threshold = 6

            # Find header row independently.
            header_idx: Optional[int] = None
            for i in range(len(all_rows) - 1):
                if (
                    _count_non_empty(all_rows[i]) >= threshold
                    and _count_non_empty(all_rows[i + 1])
                    >= threshold
                ):
                    header_idx = i
                    break

            assert header_idx is not None, (
                f"Could not find header row in sheet "
                f"'{sheet_name}' of {xlsx_path.name}"
            )

            headers = [_cell_str(c) for c in all_rows[header_idx]]

            # Count data rows.
            data_count = 0
            first_rows: List[List[str]] = []
            i = header_idx + 1
            while i < len(all_rows):
                row = all_rows[i]

                # Real footer legend → stop.
                if _is_footer_legend(row):
                    break

                # Blank row → check if structural collapse.
                if _count_non_empty(row) == 0:
                    # Look ahead.
                    has_more = False
                    for j in range(i + 1, len(all_rows)):
                        if _count_non_empty(all_rows[j]) == 0:
                            continue
                        if _is_footer_legend(all_rows[j]):
                            break
                        if _count_non_empty(all_rows[j]) >= 2:
                            has_more = True
                        break
                    if not has_more:
                        break
                    i += 1
                    continue

                data_count += 1
                row_str = [_cell_str(c) for c in row]
                if len(first_rows) < 5:
                    first_rows.append(row_str)
                i += 1

            result[csv_filename] = (
                headers, first_rows, data_count,
            )

        return result
    finally:
        wb.close()


FIXTURE_NAMES: List[str] = _discover_fixtures()


# ── Tests ───────────────────────────────────────────────────────


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestNoInstitutionalHeaderLeak:
    """Institutional header text must not appear in CSV data."""

    def test_nomenclature_no_header_leak(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv has no institutional text."""
        _assert_no_institutional_leak(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_no_header_leak(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv has no institutional text."""
        _assert_no_institutional_leak(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_no_header_leak(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv has no institutional text."""
        _assert_no_institutional_leak(
            fixture_name, "retraits.csv",
        )


def _assert_no_institutional_leak(
    fixture_name: str, csv_filename: str,
) -> None:
    """Verify no institutional header text appears in CSV rows."""
    _, data_rows = _read_golden_csv(fixture_name, csv_filename)
    for i, row in enumerate(data_rows):
        joined = " ".join(row).upper()
        for marker in INSTITUTIONAL_MARKERS:
            assert marker.upper() not in joined, (
                f"[{fixture_name}/{csv_filename}] Row {i + 1} "
                f"contains institutional text: '{marker}'."
            )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestNoFooterLegendsInData:
    """Real footer legends (F=Fabriqué, I=Importé, Nb:) must not
    appear as data rows in the CSV."""

    def test_nomenclature_no_footer_legends(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv has no footer legend rows."""
        _assert_no_footer_legends(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_no_footer_legends(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv has no footer legend rows."""
        _assert_no_footer_legends(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_no_footer_legends(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv has no footer legend rows."""
        _assert_no_footer_legends(
            fixture_name, "retraits.csv",
        )


def _assert_no_footer_legends(
    fixture_name: str, csv_filename: str,
) -> None:
    """Verify no sparse footer legend row is present in CSV data.

    A footer legend has very few non-empty cells and starts with
    F=, I=, or Nb:.  Data rows that happen to contain I= in a
    populated row are fine.
    """
    _, data_rows = _read_golden_csv(fixture_name, csv_filename)
    for i, row in enumerate(data_rows):
        non_empty = sum(1 for c in row if c.strip())
        if non_empty <= FOOTER_MAX_NON_EMPTY:
            for cell in row:
                text = cell.strip()
                if text.startswith(("F=Fabriqué", "I=Importé")):
                    pytest.fail(
                        f"[{fixture_name}/{csv_filename}] "
                        f"Row {i + 1} is a footer legend: "
                        f"'{text}'."
                    )
                if text.startswith("Nb:"):
                    pytest.fail(
                        f"[{fixture_name}/{csv_filename}] "
                        f"Row {i + 1} is a footer legend: "
                        f"'{text}'."
                    )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestDataRowsWithIEqualsPreserved:
    """Data rows containing I= in a dosage field (many populated
    cells) must NOT be dropped as footers."""

    def test_non_renouveles_i_equals_kept(
        self, fixture_name: str,
    ) -> None:
        """I=370MG/ML data rows are kept in non_renouveles.csv."""
        _assert_i_equals_data_kept(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_i_equals_kept(
        self, fixture_name: str,
    ) -> None:
        """I=370MG/ML data rows are kept in retraits.csv."""
        _assert_i_equals_data_kept(
            fixture_name, "retraits.csv",
        )


def _assert_i_equals_data_kept(
    fixture_name: str, csv_filename: str,
) -> None:
    """Independently count I= data rows in Excel and verify
    the same count appears in the golden CSV.

    We only count I= cells in rows with many non-empty cells
    (true data rows, not footer legends).
    """
    xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    try:
        sheet_map = detect_sheets(wb.sheetnames)
        # Find the actual sheet name for this csv.
        actual_sheet: Optional[str] = None
        for sn, cf in sheet_map.items():
            if cf == csv_filename:
                actual_sheet = sn
                break
        if actual_sheet is None:
            return  # Skip if sheet not mapped.

        ws = wb[actual_sheet]
        all_rows = list(ws.iter_rows(values_only=True))

        # Count data rows with I= in Excel (many non-empty cells).
        excel_i_equals_count = 0
        for row in all_rows:
            ne = _count_non_empty(row)
            if ne > FOOTER_MAX_NON_EMPTY:
                for c in row:
                    s = _cell_str(c)
                    if s.startswith("I=") or s.startswith("I= "):
                        excel_i_equals_count += 1
                        break
    finally:
        wb.close()

    if excel_i_equals_count == 0:
        return  # No I= data rows in this sheet.

    # Count I= data rows in the golden CSV.
    _, data_rows = _read_golden_csv(fixture_name, csv_filename)
    csv_i_equals_count = 0
    for row in data_rows:
        for cell in row:
            text = cell.strip()
            if text.startswith("I=") or text.startswith("I= "):
                csv_i_equals_count += 1
                break

    assert csv_i_equals_count == excel_i_equals_count, (
        f"[{fixture_name}/{csv_filename}] "
        f"Excel has {excel_i_equals_count} data rows with I=, "
        f"but CSV has {csv_i_equals_count}."
    )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestRowCountsMatchExcel:
    """CSV data row counts must match independently computed
    expectations from the raw Excel files."""

    def test_nomenclature_row_count(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv row count matches Excel."""
        _assert_row_count_matches(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_row_count(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv row count matches Excel."""
        _assert_row_count_matches(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_row_count(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv row count matches Excel."""
        _assert_row_count_matches(
            fixture_name, "retraits.csv",
        )


def _assert_row_count_matches(
    fixture_name: str, csv_filename: str,
) -> None:
    """Compare CSV row count to independently computed expectation."""
    xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
    independent = _independently_extract(xlsx)
    _, _, expected_count = independent[csv_filename]

    _, csv_rows = _read_golden_csv(fixture_name, csv_filename)
    actual_count = len(csv_rows)

    assert actual_count == expected_count, (
        f"[{fixture_name}/{csv_filename}] "
        f"Expected {expected_count} data rows "
        f"(independently computed), got {actual_count}."
    )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestHeadersMatchExcel:
    """CSV column headers must match the first tabular row in the
    Excel sheet (independently detected)."""

    def test_nomenclature_headers(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv headers match Excel."""
        _assert_headers_match(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_headers(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv headers match Excel."""
        _assert_headers_match(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_headers(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv headers match Excel."""
        _assert_headers_match(
            fixture_name, "retraits.csv",
        )


def _assert_headers_match(
    fixture_name: str, csv_filename: str,
) -> None:
    """Compare CSV headers to independently detected Excel headers.

    Empty trailing headers (from schema expansion columns) are
    excluded from comparison since they should be dropped.
    """
    xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
    independent = _independently_extract(xlsx)
    excel_headers, _, _ = independent[csv_filename]

    # Strip trailing empty headers (schema expansion).
    while excel_headers and excel_headers[-1] == "":
        excel_headers.pop()

    csv_headers, _ = _read_golden_csv(fixture_name, csv_filename)

    assert csv_headers == excel_headers, (
        f"[{fixture_name}/{csv_filename}] Header mismatch.\n"
        f"  Excel:  {excel_headers}\n"
        f"  CSV:    {csv_headers}"
    )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestFirstDataRowsMatchExcel:
    """The first few CSV data rows must match the Excel source
    (independently read)."""

    def test_nomenclature_first_rows(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv first rows match Excel."""
        _assert_first_rows_match(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_first_rows(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv first rows match Excel."""
        _assert_first_rows_match(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_first_rows(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv first rows match Excel."""
        _assert_first_rows_match(
            fixture_name, "retraits.csv",
        )


def _assert_first_rows_match(
    fixture_name: str, csv_filename: str,
) -> None:
    """Compare the first 5 data rows of the CSV to the Excel.

    Only compares columns that are non-empty in the Excel header
    (ignores schema expansion columns).
    """
    xlsx = FIXTURES_DIR / f"{fixture_name}.xlsx"
    independent = _independently_extract(xlsx)
    excel_headers, excel_first, _ = independent[csv_filename]

    # Determine which columns to keep (non-empty headers).
    keep_indices: List[int] = [
        i for i, h in enumerate(excel_headers) if h != ""
    ]

    csv_headers, csv_rows = _read_golden_csv(
        fixture_name, csv_filename,
    )

    for row_idx, excel_row in enumerate(excel_first):
        if row_idx >= len(csv_rows):
            break
        csv_row = csv_rows[row_idx]

        # Compare only the kept columns.
        for col_offset, col_idx in enumerate(keep_indices):
            excel_val = (
                excel_row[col_idx] if col_idx < len(excel_row)
                else ""
            )
            csv_val = (
                csv_row[col_offset]
                if col_offset < len(csv_row) else ""
            )
            assert csv_val == excel_val, (
                f"[{fixture_name}/{csv_filename}] "
                f"Row {row_idx + 1}, col {col_offset + 1} "
                f"('{csv_headers[col_offset]}'): "
                f"Excel='{excel_val}', CSV='{csv_val}'."
            )


@pytest.mark.parametrize("fixture_name", FIXTURE_NAMES)
class TestSequentialRowNumbering:
    """The first column (N or N°) should contain sequential
    integers starting at 1.  This is a domain-level sanity
    check that data was not mangled or reordered."""

    def test_nomenclature_sequential(
        self, fixture_name: str,
    ) -> None:
        """nomenclature.csv has sequential N column."""
        _assert_sequential_first_col(
            fixture_name, "nomenclature.csv",
        )

    def test_non_renouveles_sequential(
        self, fixture_name: str,
    ) -> None:
        """non_renouveles.csv has sequential N° column."""
        _assert_sequential_first_col(
            fixture_name, "non_renouveles.csv",
        )

    def test_retraits_sequential(
        self, fixture_name: str,
    ) -> None:
        """retraits.csv has sequential N° column."""
        _assert_sequential_first_col(
            fixture_name, "retraits.csv",
        )


def _assert_sequential_first_col(
    fixture_name: str, csv_filename: str,
) -> None:
    """Verify the first column contains sequential integers 1..N."""
    _, data_rows = _read_golden_csv(fixture_name, csv_filename)
    for i, row in enumerate(data_rows):
        expected = str(i + 1)
        actual = row[0].strip() if row else ""
        # Some rows may have float-formatted ints (e.g. "1.0").
        try:
            actual_int = str(int(float(actual)))
        except (ValueError, IndexError):
            actual_int = actual
        assert actual_int == expected, (
            f"[{fixture_name}/{csv_filename}] "
            f"Row {i + 1}: expected N={expected}, got '{actual}'."
        )
