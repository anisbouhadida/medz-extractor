"""CLI entry point for medz-extractor.

Command:
    medz-extractor process <input.xlsx> --out <output_dir>
"""

import logging
import time
from pathlib import Path

import typer
from openpyxl import load_workbook

from medz_extractor.parser import parse_sheet
from medz_extractor.schema import drop_empty_columns
from medz_extractor.sheet_detector import detect_sheets
from medz_extractor.writer import write_csv

app = typer.Typer(
    help=(
        "Convert official Algerian pharmaceutical nomenclature "
        "Excel reports (.xlsx) into clean CSV datasets.\n\n"
        "Produces exactly 3 files per workbook:\n\n"
        "  nomenclature.csv  non_renouveles.csv  retraits.csv"
    ),
    rich_markup_mode="rich",
    no_args_is_help=True,
)

# Configure root logger for the package.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)


@app.command()
def process(
    input_file: Path = typer.Argument(
        ...,
        help="Path to the input .xlsx nomenclature workbook.",
        exists=True,
        readable=True,
        resolve_path=True,
    ),
    out: Path = typer.Option(
        ...,
        "--out",
        help=(
            "Directory where the 3 CSV files will be written "
            "(created automatically if it does not exist)."
        ),
        resolve_path=True,
    ),
    delimiter: str = typer.Option(
        ",",
        "--delimiter",
        help="CSV field delimiter (e.g. ',' or ';').",
    ),
    dry_run: bool = typer.Option(
        False,
        "--dry-run",
        help=(
            "Run the full parse pipeline (sheet detection, "
            "header detection, row extraction) and report "
            "results without writing any CSV files.  "
            "Useful for validating a new .xlsx before processing."
        ),
    ),
) -> None:
    """Process an Excel nomenclature workbook into clean CSVs.

    Reads the workbook, detects the 3 expected sheets
    (Nomenclature, Non Renouveles, Retraits), extracts tabular
    data, drops empty columns, and writes normalised CSV outputs.

    Use --dry-run to inspect detected sheets and row counts
    without writing any files.

    [bold]Examples:[/bold]

      medz-extractor input/2025-11.xlsx --out output/2025-11/

      medz-extractor input/2025-11.xlsx --out output/2025-11/ --delimiter ';'

      medz-extractor input/2025-11.xlsx --out output/2025-11/ --dry-run
    """
    start = time.monotonic()
    logger.info("Input file: %s", input_file)
    logger.info("Output directory: %s", out)

    # --- Load workbook (read-only, no macros). ---
    try:
        wb = load_workbook(str(input_file), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Failed to open workbook: %s", exc)
        raise typer.Exit(code=1) from exc

    # --- Detect required sheets. ---
    try:
        sheet_map = detect_sheets(wb.sheetnames)
    except ValueError as exc:
        logger.error("%s", exc)
        wb.close()
        raise typer.Exit(code=1) from exc

    logger.info("Detected sheets: %s", sheet_map)

    if dry_run:
        logger.info("Dry-run mode — no CSV files will be written.")

    # --- Process each sheet. ---
    generated: list[Path] = []
    try:
        for sheet_name, csv_filename in sheet_map.items():
            logger.info(
                "Processing sheet '%s' → '%s' …",
                sheet_name,
                csv_filename,
            )
            ws = wb[sheet_name]

            # Parse: header detection + data extraction.
            try:
                headers, data_rows = parse_sheet(ws, csv_filename)
            except ValueError as exc:
                logger.error(
                    "Parsing failed for sheet '%s': %s",
                    sheet_name,
                    exc,
                )
                raise typer.Exit(code=1) from exc

            logger.info(
                "Sheet '%s': header row found, %d data rows extracted.",
                sheet_name,
                len(data_rows),
            )

            # Schema normalization: drop empty columns.
            headers, data_rows = drop_empty_columns(headers, data_rows)

            if dry_run:
                logger.info(
                    "Sheet '%s': %d columns, %d rows "
                    "(skipping CSV write).",
                    sheet_name,
                    len(headers),
                    len(data_rows),
                )
                continue

            # Write CSV.
            csv_path = out / csv_filename
            try:
                write_csv(headers, data_rows, csv_path, delimiter)
            except (OSError, ValueError) as exc:
                logger.error(
                    "CSV write failed for '%s': %s",
                    csv_path,
                    exc,
                )
                raise typer.Exit(code=1) from exc

            generated.append(csv_path)
    finally:
        wb.close()

    elapsed = time.monotonic() - start
    if dry_run:
        logger.info("Dry-run finished in %.2f s. No files written.", elapsed)
    else:
        logger.info(
            "Done. Generated %d CSV file(s) in %.2f s:",
            len(generated),
            elapsed,
        )
        for p in generated:
            logger.info("  %s", p)
